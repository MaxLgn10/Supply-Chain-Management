"""
fill_solution_template.py

Reads the metaheuristic optimization output CSVs and writes the solution
into the official PPP_solutionFile_2026.xlsx template.

Fills:
  - Packs sheet        (row 3 = SKU headers, rows 4+ = units per SKU per pack)
  - PackAllocation sheet (row 1 = channel headers, rows 2+ = packs per channel)

Run from project root:
  python fill_solution_template.py                          # uses optimization Wout/outputs/
  python fill_solution_template.py "feature tests/outputs"  # uses a different outputs folder
"""

import sys
from pathlib import Path

import openpyxl
import pandas as pd

# Allow passing a custom outputs folder as CLI argument
_outputs_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("optimization Wout/outputs")

PACKS_CSV    = _outputs_dir / "optimalisation_packs_metaheuristic.csv"
ALLOC_CSV    = _outputs_dir / "optimalisation_pack_allocation_metaheuristic.csv"
TEMPLATE     = Path("data/PPP_solutionFile_2026.xlsx")
OUTPUT = Path("outputs/PPP_solutionFile_2026_filled.xlsx")

# ---------------------------------------------------------------------------

def main() -> None:
    for p in (PACKS_CSV, ALLOC_CSV, TEMPLATE):
        if not p.exists():
            sys.exit(f"ERROR: file not found: {p}")

    OUTPUT.parent.mkdir(exist_ok=True)

    # ── Load CSV data ────────────────────────────────────────────────────────
    df_packs = pd.read_csv(PACKS_CSV, index_col=0)   # index = P001, P002, ...
    df_alloc  = pd.read_csv(ALLOC_CSV,  index_col=0)

    # drop the human-readable name column; only keep SKU quantity columns
    if "pack_name" in df_packs.columns:
        df_packs = df_packs.drop(columns=["pack_name"])

    n_packs = len(df_packs)
    print(f"Packs to write     : {n_packs}")
    print(f"SKUs in solution   : {len(df_packs.columns)}")
    print(f"Channels in alloc  : {list(df_alloc.columns)}")

    # ── Open template ────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(TEMPLATE)

    # ── Fill Packs sheet ─────────────────────────────────────────────────────
    ws_packs = wb["Packs"]

    # Read the SKU column order from template row 3 (1-indexed, col A is index)
    sku_to_col: dict[str, int] = {}
    for cell in ws_packs[3]:
        if cell.value and cell.column > 1:        # skip col A
            sku_to_col[cell.value] = cell.column

    # Verify all our SKUs are present in the template
    missing = set(df_packs.columns) - set(sku_to_col)
    if missing:
        sys.exit(f"ERROR: {len(missing)} SKUs in solution not found in template: {list(missing)[:5]}")

    # Clear any existing data rows (4 to 1004)
    for row_idx in range(4, 1005):
        for col_idx in range(1, ws_packs.max_column + 1):
            ws_packs.cell(row=row_idx, column=col_idx).value = None

    # Write solution
    for pack_num, (pack_id, pack_row) in enumerate(df_packs.iterrows()):
        row = 4 + pack_num          # row 4 = pack index 0
        ws_packs.cell(row=row, column=1, value=pack_num)   # col A = 0-based index
        for sku, col in sku_to_col.items():
            val = int(pack_row.get(sku, 0))
            if val != 0:
                ws_packs.cell(row=row, column=col, value=val)

    print(f"Packs sheet        : wrote {n_packs} rows (rows 4–{3+n_packs})")

    # ── Fill PackAllocation sheet ────────────────────────────────────────────
    ws_alloc = wb["PackAllocation"]

    # Read channel column order from template row 1
    ch_to_col: dict[str, int] = {}
    for cell in ws_alloc[1]:
        if cell.value and cell.column > 1 and cell.column <= 12:   # cols B–L = 11 channels
            ch_to_col[cell.value] = cell.column

    # Verify channels
    missing_ch = set(df_alloc.columns) - set(ch_to_col)
    if missing_ch:
        sys.exit(f"ERROR: channels in solution not found in template: {missing_ch}")

    # Clear data rows 2–1002, cols A–L only (preserve cols M+ = formulas)
    for row_idx in range(2, 1003):
        for col_idx in range(1, 13):    # cols A–L
            ws_alloc.cell(row=row_idx, column=col_idx).value = None

    # Write solution
    for pack_num, (pack_id, alloc_row) in enumerate(df_alloc.iterrows()):
        row = 2 + pack_num          # row 2 = pack index 0
        ws_alloc.cell(row=row, column=1, value=pack_num)   # col A = 0-based index
        for channel, col in ch_to_col.items():
            val = int(alloc_row.get(channel, 0))
            if val != 0:
                ws_alloc.cell(row=row, column=col, value=val)

    print(f"PackAllocation sheet: wrote {n_packs} rows (rows 2–{1+n_packs})")

    # ── Save ─────────────────────────────────────────────────────────────────
    wb.save(OUTPUT)
    print(f"\nSaved -> {OUTPUT}")

    # ── Quick sanity check ───────────────────────────────────────────────────
    total_alloc = int(df_alloc.sum().sum())
    print(f"Total packs allocated: {total_alloc:,}  (expected 4,708)")


if __name__ == "__main__":
    main()
