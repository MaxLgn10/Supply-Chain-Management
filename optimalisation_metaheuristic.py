#!/usr/bin/env python3
"""
optimalisation.py

Metaheuristic pre-pack optimizer for the SCM group assignment.

This version does NOT fix pack contents first and then solve allocation separately.
It optimizes the three decisions together in one metaheuristic state:

    1. how many pack types are used
    2. what SKU quantities each pack contains
    3. how many packs of each type are allocated to each sales channel

The implemented method is a hybrid metaheuristic inspired by common approaches
used for the PrePack Optimization Problem (POP):

    - demand-driven size-curve pack construction
    - greedy randomized construction
    - adaptive large-neighborhood search style destroy/repair moves
    - simulated annealing acceptance to escape local minima
    - local allocation repair and pruning after each content change

The model minimizes:

    pack creation cost
    + pack handling cost
    + capital cost of overstock
    + shortage penalty

Main input:
    outputs/master_forecast_2026.csv

Required columns:
    channel_id, product_id, size, forecast_ensemble

Optional columns:
    category_group

Optional product file:
    data/PPP_stu_products.xlsx
    Used to read SKU unit cost. If missing, all unit costs are set to 1.

Optional official solution template:
    data/PPP_solutionFile_2026.xlsx
    If present, the script tries to fill the Packs and PackAllocation sheets.
    If not present, it creates a clean workbook with those two sheets.

Outputs:
    outputs/optimalisation_solution.xlsx
    outputs/optimalisation_packs.csv
    outputs/optimalisation_pack_allocation.csv
    outputs/optimalisation_assortment.csv
    outputs/optimalisation_diagnostics.csv
    outputs/optimalisation_cost_summary.csv
    outputs/optimalisation_run_log.csv

Run:
    python optimalisation.py

Recommended stronger run:
    python optimalisation.py --iterations 50000 --restarts 8 --time-limit 1800

The file name "optimalisation.py" is kept because that is what was requested.
"""

from __future__ import annotations

import argparse
import copy
import math
import random
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd


# =============================================================================
# Cost parameters from the current assignment code base
# =============================================================================

COST_OF_CAPITAL = 0.243
HANDLING_COST_PER_PACK = 11.03
PACK_CREATION_COST = 134.00

# A very high shortage penalty keeps shortages unattractive but still allows
# the metaheuristic to pass through temporarily infeasible intermediate states.
SHORTAGE_PENALTY_PER_UNIT = 10_000.00

# Operational limits. Tune these if the official template or teacher imposes
# stricter requirements.
MAX_PACK_UNITS = 10000
MAX_DISTINCT_SKUS_PER_PACK = 10000
MAX_PACK_TYPES_DEFAULT = None  # None = no hard limit unless passed by CLI.

# Allocation repair limits.
MAX_REPAIR_STEPS_PER_CHANNEL = 20_000
PRUNE_PASSES = 3


# =============================================================================
# Data classes
# =============================================================================

@dataclass
class ProblemData:
    demand: np.ndarray
    sku_ids: List[str]
    channel_ids: List[str]
    unit_cost: np.ndarray
    sku_meta: pd.DataFrame
    sku_index: Dict[str, int]
    channel_index: Dict[str, int]


@dataclass
class Solution:
    # content[p, s] = units of SKU s inside pack p
    content: np.ndarray

    # alloc[p, c] = number of packs p sent to channel c
    alloc: np.ndarray

    # Active pack type names. Length equals number of rows in content/alloc.
    names: List[str]

    # Cached values; recalculated by evaluate().
    cost: Optional[float] = None
    setup_cost: Optional[float] = None
    handling_cost: Optional[float] = None
    capital_cost: Optional[float] = None
    shortage_cost: Optional[float] = None
    overstock_units: Optional[int] = None
    shortage_units: Optional[int] = None
    shipped_units: Optional[int] = None
    allocated_packs: Optional[int] = None


# =============================================================================
# Utility functions
# =============================================================================

def seed_everything(seed: int) -> None:
    random.seed(seed)
    np.random.seed(seed)


def make_sku_id(product_id, size) -> str:
    return f"{product_id}_{size}"


def slugify(value) -> str:
    value = str(value).lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_")


def ceil_nonnegative(value) -> int:
    if pd.isna(value):
        return 0
    return max(0, int(math.ceil(float(value))))


def safe_int_matrix(a: np.ndarray) -> np.ndarray:
    return np.maximum(0, np.rint(a).astype(int))


def proportional_integer_vector(
    weights: np.ndarray,
    total_units: int,
    max_distinct: int,
    rng: random.Random,
) -> np.ndarray:
    """
    Build an integer vector with approximately the same proportions as weights.
    Only the largest positive weights are kept.
    """
    n = len(weights)
    result = np.zeros(n, dtype=int)

    if total_units <= 0:
        return result

    positive_idx = np.flatnonzero(weights > 0)
    if len(positive_idx) == 0:
        return result

    # Keep the most relevant SKUs, with a small randomization among near-ties.
    scored = [(idx, float(weights[idx]) * rng.uniform(0.90, 1.10)) for idx in positive_idx]
    scored.sort(key=lambda x: x[1], reverse=True)

    keep_count = min(max_distinct, total_units, len(scored))
    kept = [idx for idx, _ in scored[:keep_count]]
    kept_weights = weights[kept].astype(float)

    if kept_weights.sum() <= 0:
        return result

    raw = kept_weights / kept_weights.sum() * total_units
    base = np.floor(raw).astype(int)

    # Give every selected SKU at least 1 unit when possible.
    if total_units >= keep_count:
        base = np.maximum(base, 1)

    diff = total_units - int(base.sum())

    if diff > 0:
        remainders = raw - np.floor(raw)
        order = np.argsort(-remainders)
        for k in order[:diff]:
            base[k] += 1
    elif diff < 0:
        # Remove from largest quantities while keeping positive selected SKUs.
        order = np.argsort(-base)
        for k in order:
            while diff < 0 and base[k] > 1:
                base[k] -= 1
                diff += 1
            if diff == 0:
                break

    for idx, qty in zip(kept, base):
        if qty > 0:
            result[idx] = int(qty)

    return result


def remove_empty_pack_rows(sol: Solution) -> Solution:
    if sol.content.size == 0:
        return sol

    used_content = sol.content.sum(axis=1) > 0
    used_alloc = sol.alloc.sum(axis=1) > 0
    keep = used_content | used_alloc

    if keep.all():
        return sol

    if not keep.any():
        n_skus = sol.content.shape[1]
        n_channels = sol.alloc.shape[1]
        return Solution(
            content=np.zeros((0, n_skus), dtype=int),
            alloc=np.zeros((0, n_channels), dtype=int),
            names=[],
        )

    return Solution(
        content=sol.content[keep].copy(),
        alloc=sol.alloc[keep].copy(),
        names=[name for name, k in zip(sol.names, keep) if k],
    )


def ensure_unique_names(names: List[str]) -> List[str]:
    seen = {}
    out = []
    for name in names:
        base = str(name) if str(name) else "pack"
        seen[base] = seen.get(base, 0) + 1
        if seen[base] == 1:
            out.append(base)
        else:
            out.append(f"{base}_{seen[base]}")
    return out


# =============================================================================
# Loading data
# =============================================================================

def load_problem(
    forecast_path: str,
    product_path: str,
    min_forecast: int = 1,
) -> ProblemData:
    forecast_file = Path(forecast_path)
    if not forecast_file.exists():
        raise FileNotFoundError(
            f"Forecast file not found: {forecast_path}. "
            "Run run_forecasting.py first to create outputs/master_forecast_2026.csv."
        )

    df = pd.read_csv(forecast_file)

    required = {"channel_id", "product_id", "size", "forecast_ensemble"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Forecast file is missing required columns: {sorted(missing)}")

    df = df.copy()
    df["sku_id"] = [make_sku_id(pid, size) for pid, size in zip(df["product_id"], df["size"])]
    df["channel_id"] = df["channel_id"].astype(str)
    df["forecast_units"] = df["forecast_ensemble"].apply(ceil_nonnegative)

    group_cols = ["sku_id", "channel_id", "product_id", "size"]
    if "category_group" in df.columns:
        group_cols.append("category_group")

    df = (
        df.groupby(group_cols, dropna=False)["forecast_units"]
          .sum()
          .reset_index()
    )

    demand_df = (
        df.pivot_table(
            index="sku_id",
            columns="channel_id",
            values="forecast_units",
            aggfunc="sum",
            fill_value=0,
        )
        .astype(int)
        .sort_index()
    )

    demand_df = demand_df[demand_df.sum(axis=1) >= min_forecast]

    if demand_df.empty:
        raise ValueError("No SKU has positive forecast demand after filtering.")

    sku_ids = list(demand_df.index)
    channel_ids = list(demand_df.columns)

    sku_meta_cols = ["sku_id", "product_id", "size"]
    if "category_group" in df.columns:
        sku_meta_cols.append("category_group")

    sku_meta = (
        df[sku_meta_cols]
        .drop_duplicates("sku_id")
        .set_index("sku_id")
        .reindex(sku_ids)
        .reset_index()
    )

    # Unit cost
    product_file = Path(product_path)
    if product_file.exists():
        products = pd.read_excel(product_file)
        if {"id", "cost"}.issubset(products.columns):
            cost_by_product = products.set_index("id")["cost"].to_dict()
            unit_cost = np.array(
                [float(cost_by_product.get(pid, 1.0)) for pid in sku_meta["product_id"]],
                dtype=float,
            )
        else:
            print("WARNING: product file has no id/cost columns; using unit cost = 1.")
            unit_cost = np.ones(len(sku_ids), dtype=float)
    else:
        print(f"WARNING: {product_path} not found; using unit cost = 1.")
        unit_cost = np.ones(len(sku_ids), dtype=float)

    return ProblemData(
        demand=demand_df.to_numpy(dtype=int),
        sku_ids=sku_ids,
        channel_ids=channel_ids,
        unit_cost=unit_cost,
        sku_meta=sku_meta,
        sku_index={sku: i for i, sku in enumerate(sku_ids)},
        channel_index={ch: i for i, ch in enumerate(channel_ids)},
    )


# =============================================================================
# Cost evaluation
# =============================================================================

def evaluate(sol: Solution, data: ProblemData) -> float:
    if sol.content.shape[0] == 0:
        shipped = np.zeros_like(data.demand)
        allocated_packs = 0
        active_pack_types = 0
    else:
        shipped = sol.content.T @ sol.alloc
        allocated_packs = int(sol.alloc.sum())
        active_pack_types = int(((sol.content.sum(axis=1) > 0) & (sol.alloc.sum(axis=1) > 0)).sum())

    over = np.maximum(0, shipped - data.demand)
    under = np.maximum(0, data.demand - shipped)

    setup_cost = active_pack_types * PACK_CREATION_COST
    handling_cost = allocated_packs * HANDLING_COST_PER_PACK
    capital_cost = float((over * data.unit_cost[:, None] * COST_OF_CAPITAL).sum())
    shortage_cost = float(under.sum() * SHORTAGE_PENALTY_PER_UNIT)

    total = setup_cost + handling_cost + capital_cost + shortage_cost

    sol.cost = float(total)
    sol.setup_cost = float(setup_cost)
    sol.handling_cost = float(handling_cost)
    sol.capital_cost = float(capital_cost)
    sol.shortage_cost = float(shortage_cost)
    sol.overstock_units = int(over.sum())
    sol.shortage_units = int(under.sum())
    sol.shipped_units = int(shipped.sum())
    sol.allocated_packs = allocated_packs

    return float(total)


def shipped_matrix(sol: Solution, data: ProblemData) -> np.ndarray:
    if sol.content.shape[0] == 0:
        return np.zeros_like(data.demand)
    return sol.content.T @ sol.alloc


def delta_add_one_pack(
    content_vec: np.ndarray,
    shipped_col: np.ndarray,
    demand_col: np.ndarray,
    unit_cost: np.ndarray,
) -> float:
    """
    Cost difference of adding one pack with content_vec to a specific channel.
    """
    before_over = np.maximum(0, shipped_col - demand_col)
    before_under = np.maximum(0, demand_col - shipped_col)

    after = shipped_col + content_vec
    after_over = np.maximum(0, after - demand_col)
    after_under = np.maximum(0, demand_col - after)

    delta_capital = float(((after_over - before_over) * unit_cost * COST_OF_CAPITAL).sum())
    delta_shortage = float((after_under.sum() - before_under.sum()) * SHORTAGE_PENALTY_PER_UNIT)
    return HANDLING_COST_PER_PACK + delta_capital + delta_shortage


def delta_remove_one_pack(
    content_vec: np.ndarray,
    shipped_col: np.ndarray,
    demand_col: np.ndarray,
    unit_cost: np.ndarray,
) -> float:
    """
    Cost difference of removing one pack with content_vec from a specific channel.
    """
    before_over = np.maximum(0, shipped_col - demand_col)
    before_under = np.maximum(0, demand_col - shipped_col)

    after = shipped_col - content_vec
    after_over = np.maximum(0, after - demand_col)
    after_under = np.maximum(0, demand_col - after)

    delta_capital = float(((after_over - before_over) * unit_cost * COST_OF_CAPITAL).sum())
    delta_shortage = float((after_under.sum() - before_under.sum()) * SHORTAGE_PENALTY_PER_UNIT)
    return -HANDLING_COST_PER_PACK + delta_capital + delta_shortage


# =============================================================================
# Construction and repair
# =============================================================================

def create_single_sku_initial_solution(data: ProblemData) -> Solution:
    """
    Feasible but expensive baseline:
        one pack type per SKU, one unit per pack, exact allocation to demand.
    """
    n_skus, n_channels = data.demand.shape

    content = np.eye(n_skus, dtype=int)
    alloc = data.demand.copy().T  # channels x skus
    alloc = alloc.T              # skus x channels

    names = [f"single_{slugify(sku)}" for sku in data.sku_ids]

    sol = Solution(content=content, alloc=alloc, names=names)
    evaluate(sol, data)
    return sol


def make_curve_pack_from_indices(
    indices: Sequence[int],
    weights: np.ndarray,
    rng: random.Random,
    min_units: int = 4,
    max_units: int = MAX_PACK_UNITS,
) -> np.ndarray:
    if not indices:
        return np.zeros_like(weights, dtype=int)

    indices = list(indices)
    total = rng.choice([4, 6, 8, 10, 12, 16, 20, 24])
    total = max(min_units, min(max_units, total))

    local_weights = np.zeros_like(weights, dtype=float)
    local_weights[indices] = weights[indices]

    pack = proportional_integer_vector(
        local_weights,
        total_units=total,
        max_distinct=min(MAX_DISTINCT_SKUS_PER_PACK, len(indices)),
        rng=rng,
    )

    return pack


def create_random_curve_pack(data: ProblemData, rng: random.Random) -> Tuple[np.ndarray, str]:
    """
    Create a demand-driven candidate pack:
        - same product size curve
        - same category curve
        - channel-specific curve
        - global high-demand curve
    """
    demand_total = data.demand.sum(axis=1).astype(float)
    n_skus = len(data.sku_ids)

    mode_options = ["product", "channel_product", "category", "channel", "global"]
    mode = rng.choice(mode_options)

    meta = data.sku_meta.copy()

    if mode == "product":
        pid = rng.choice(list(meta["product_id"].dropna().unique()))
        idx = meta.index[meta["product_id"] == pid].tolist()
        pack = make_curve_pack_from_indices(idx, demand_total, rng)
        return pack, f"prod_{pid}_curve"

    if mode == "channel_product":
        ch_idx = rng.randrange(len(data.channel_ids))
        ch = data.channel_ids[ch_idx]
        positive_products = (
            meta.assign(demand=data.demand[:, ch_idx])
                .query("demand > 0")["product_id"]
                .dropna()
                .unique()
        )
        if len(positive_products) > 0:
            pid = rng.choice(list(positive_products))
            idx = meta.index[(meta["product_id"] == pid) & (data.demand[:, ch_idx] > 0)].tolist()
            pack = make_curve_pack_from_indices(idx, data.demand[:, ch_idx].astype(float), rng)
            return pack, f"prod_{pid}_{slugify(ch)}_curve"

    if mode == "category" and "category_group" in meta.columns:
        categories = list(meta["category_group"].dropna().unique())
        if categories:
            cat = rng.choice(categories)
            idx = meta.index[meta["category_group"] == cat].tolist()
            pack = make_curve_pack_from_indices(idx, demand_total, rng)
            return pack, f"cat_{slugify(cat)}_curve"

    if mode == "channel":
        ch_idx = rng.randrange(len(data.channel_ids))
        ch = data.channel_ids[ch_idx]
        idx = np.flatnonzero(data.demand[:, ch_idx] > 0).tolist()
        pack = make_curve_pack_from_indices(idx, data.demand[:, ch_idx].astype(float), rng)
        return pack, f"channel_{slugify(ch)}_curve"

    # Global fallback
    idx = np.flatnonzero(demand_total > 0).tolist()
    pack = make_curve_pack_from_indices(idx, demand_total, rng)
    return pack, "global_curve"


def append_pack(sol: Solution, pack: np.ndarray, name: str) -> Solution:
    if pack.sum() <= 0:
        return sol

    pack = pack.astype(int)
    pack = np.maximum(0, pack)

    if pack.sum() > MAX_PACK_UNITS:
        # Scale down by removing from largest quantities.
        while pack.sum() > MAX_PACK_UNITS:
            j = int(np.argmax(pack))
            pack[j] -= 1

    if np.count_nonzero(pack) > MAX_DISTINCT_SKUS_PER_PACK:
        keep = np.argsort(-pack)[:MAX_DISTINCT_SKUS_PER_PACK]
        new_pack = np.zeros_like(pack)
        new_pack[keep] = pack[keep]
        pack = new_pack

    # Do not add exact duplicates.
    if sol.content.shape[0] > 0:
        duplicate = np.any(np.all(sol.content == pack[None, :], axis=1))
        if duplicate:
            return sol

    n_channels = sol.alloc.shape[1] if sol.alloc.size else 0
    if n_channels == 0:
        raise ValueError("Solution has no channel dimension.")

    sol.content = np.vstack([sol.content, pack[None, :]])
    sol.alloc = np.vstack([sol.alloc, np.zeros((1, n_channels), dtype=int)])
    sol.names.append(name)
    sol.names = ensure_unique_names(sol.names)
    return sol


def repair_allocation(sol: Solution, data: ProblemData, rng: random.Random) -> Solution:
    """
    Given current pack contents and allocations, improve allocations with a
    greedy randomized add/prune procedure.

    This keeps the state integrated: pack contents and allocations both evolve,
    but allocation is locally repaired after destructive content changes.
    """
    if sol.content.shape[0] == 0:
        return sol

    sol.alloc = safe_int_matrix(sol.alloc)
    sol.content = safe_int_matrix(sol.content)

    n_packs, n_skus = sol.content.shape
    n_channels = len(data.channel_ids)

    # Remove invalid empty pack types.
    nonempty = sol.content.sum(axis=1) > 0
    sol.content = sol.content[nonempty]
    sol.alloc = sol.alloc[nonempty]
    sol.names = [name for name, keep in zip(sol.names, nonempty) if keep]

    if sol.content.shape[0] == 0:
        return sol

    # Add packs while doing so reduces objective.
    shipped = shipped_matrix(sol, data)

    for c_idx in range(n_channels):
        steps = 0
        while steps < MAX_REPAIR_STEPS_PER_CHANNEL:
            shortage = data.demand[:, c_idx] - shipped[:, c_idx]
            if shortage.max() <= 0:
                break

            best = []
            for p_idx in range(sol.content.shape[0]):
                pack = sol.content[p_idx]
                if pack.sum() <= 0:
                    continue

                # Pack must cover at least one currently short SKU.
                if np.minimum(pack, np.maximum(0, shortage)).sum() <= 0:
                    continue

                delta = delta_add_one_pack(
                    content_vec=pack,
                    shipped_col=shipped[:, c_idx],
                    demand_col=data.demand[:, c_idx],
                    unit_cost=data.unit_cost,
                )

                # Stronger preference for packs covering many shortage units.
                covered_shortage = int(np.minimum(pack, np.maximum(0, shortage)).sum())
                score = delta - rng.random() * 0.01 * covered_shortage
                best.append((score, delta, p_idx))

            if not best:
                break

            best.sort(key=lambda x: x[0])

            # Randomized greedy: pick from the best few.
            shortlist = best[: min(5, len(best))]
            _, delta, chosen_p = rng.choice(shortlist)

            # Only add if it improves the penalized objective.
            if delta >= 0:
                break

            sol.alloc[chosen_p, c_idx] += 1
            shipped[:, c_idx] += sol.content[chosen_p]
            steps += 1

    # Prune allocations that became unnecessary or damaging.
    for _ in range(PRUNE_PASSES):
        improved = False
        shipped = shipped_matrix(sol, data)

        pack_channel_pairs = [
            (p_idx, c_idx)
            for p_idx in range(sol.content.shape[0])
            for c_idx in range(n_channels)
            if sol.alloc[p_idx, c_idx] > 0
        ]
        rng.shuffle(pack_channel_pairs)

        for p_idx, c_idx in pack_channel_pairs:
            while sol.alloc[p_idx, c_idx] > 0:
                delta = delta_remove_one_pack(
                    content_vec=sol.content[p_idx],
                    shipped_col=shipped[:, c_idx],
                    demand_col=data.demand[:, c_idx],
                    unit_cost=data.unit_cost,
                )
                if delta < -1e-9:
                    sol.alloc[p_idx, c_idx] -= 1
                    shipped[:, c_idx] -= sol.content[p_idx]
                    improved = True
                else:
                    break

        if not improved:
            break

    sol = remove_empty_pack_rows(sol)
    evaluate(sol, data)
    return sol


def greedy_construct_solution(data: ProblemData, rng: random.Random, n_seed_packs: int) -> Solution:
    """
    Build a stronger initial solution:
        1. start from exact single-SKU baseline,
        2. add demand curve packs,
        3. repair and prune allocation.
    """
    sol = create_single_sku_initial_solution(data)

    for _ in range(n_seed_packs):
        pack, name = create_random_curve_pack(data, rng)
        sol = append_pack(sol, pack, name)

    sol = repair_allocation(sol, data, rng)
    evaluate(sol, data)
    return sol


# =============================================================================
# Neighborhood operators
# =============================================================================

def op_add_curve_pack(sol: Solution, data: ProblemData, rng: random.Random) -> Solution:
    new = copy.deepcopy(sol)
    pack, name = create_random_curve_pack(data, rng)
    append_pack(new, pack, name)
    return repair_allocation(new, data, rng)


def op_remove_pack(sol: Solution, data: ProblemData, rng: random.Random) -> Solution:
    if sol.content.shape[0] <= 1:
        return sol

    new = copy.deepcopy(sol)

    # Prefer removing low-use or expensive pack types.
    usage = new.alloc.sum(axis=1)
    if rng.random() < 0.70:
        candidates = np.argsort(usage)[: max(1, len(usage) // 4)]
        idx = int(rng.choice(list(candidates)))
    else:
        idx = rng.randrange(new.content.shape[0])

    keep = np.ones(new.content.shape[0], dtype=bool)
    keep[idx] = False

    new.content = new.content[keep]
    new.alloc = new.alloc[keep]
    new.names = [name for i, name in enumerate(new.names) if keep[i]]

    return repair_allocation(new, data, rng)


def op_mutate_pack_content(sol: Solution, data: ProblemData, rng: random.Random) -> Solution:
    if sol.content.shape[0] == 0:
        return sol

    new = copy.deepcopy(sol)
    p = rng.randrange(new.content.shape[0])
    pack = new.content[p].copy()

    demand_total = data.demand.sum(axis=1)
    positive_skus = np.flatnonzero(demand_total > 0)

    action = rng.choice(["add", "remove", "increment", "decrement", "swap"])

    if action == "add":
        if np.count_nonzero(pack) < MAX_DISTINCT_SKUS_PER_PACK and pack.sum() < MAX_PACK_UNITS:
            # Add a SKU related to an existing product/category where possible.
            if pack.sum() > 0 and rng.random() < 0.60:
                existing = np.flatnonzero(pack > 0)
                ref = int(rng.choice(list(existing)))
                ref_product = data.sku_meta.loc[ref, "product_id"]
                related = data.sku_meta.index[data.sku_meta["product_id"] == ref_product].tolist()
                related = [i for i in related if demand_total[i] > 0 and pack[i] == 0]
                if related:
                    s = int(rng.choice(related))
                else:
                    s = int(rng.choice(list(positive_skus)))
            else:
                probs = demand_total[positive_skus].astype(float)
                probs = probs / probs.sum()
                s = int(rng.choices(list(positive_skus), weights=list(probs), k=1)[0])
            pack[s] += 1

    elif action == "remove":
        nonzero = np.flatnonzero(pack > 0)
        if len(nonzero) > 1:
            s = int(rng.choice(list(nonzero)))
            pack[s] = 0

    elif action == "increment":
        nonzero = np.flatnonzero(pack > 0)
        if len(nonzero) > 0 and pack.sum() < MAX_PACK_UNITS:
            s = int(rng.choice(list(nonzero)))
            pack[s] += 1

    elif action == "decrement":
        nonzero = np.flatnonzero(pack > 0)
        if len(nonzero) > 0:
            s = int(rng.choice(list(nonzero)))
            pack[s] = max(0, pack[s] - 1)

    elif action == "swap":
        nonzero = np.flatnonzero(pack > 0)
        zero_positive = [i for i in positive_skus if pack[i] == 0]
        if len(nonzero) > 0 and zero_positive:
            s_out = int(rng.choice(list(nonzero)))
            s_in = int(rng.choice(zero_positive))
            qty = pack[s_out]
            pack[s_out] = 0
            pack[s_in] = max(1, qty)

    if pack.sum() <= 0:
        return sol

    while pack.sum() > MAX_PACK_UNITS:
        j = int(np.argmax(pack))
        pack[j] -= 1

    if np.count_nonzero(pack) > MAX_DISTINCT_SKUS_PER_PACK:
        keep = np.argsort(-pack)[:MAX_DISTINCT_SKUS_PER_PACK]
        clipped = np.zeros_like(pack)
        clipped[keep] = pack[keep]
        pack = clipped

    new.content[p] = pack
    new.names[p] = f"mut_{new.names[p]}"

    return repair_allocation(new, data, rng)


def op_merge_packs(sol: Solution, data: ProblemData, rng: random.Random) -> Solution:
    if sol.content.shape[0] < 2:
        return sol

    new = copy.deepcopy(sol)

    p1, p2 = rng.sample(range(new.content.shape[0]), 2)
    pack = new.content[p1] + new.content[p2]

    # Clip to operational limits.
    if np.count_nonzero(pack) > MAX_DISTINCT_SKUS_PER_PACK:
        keep = np.argsort(-pack)[:MAX_DISTINCT_SKUS_PER_PACK]
        clipped = np.zeros_like(pack)
        clipped[keep] = pack[keep]
        pack = clipped

    while pack.sum() > MAX_PACK_UNITS:
        j = int(np.argmax(pack))
        pack[j] -= 1

    append_pack(new, pack, f"merge_{p1}_{p2}")

    # Sometimes remove one of the source packs to force structural change.
    if rng.random() < 0.50 and new.content.shape[0] > 2:
        remove_idx = rng.choice([p1, p2])
        keep = np.ones(new.content.shape[0], dtype=bool)
        keep[remove_idx] = False
        new.content = new.content[keep]
        new.alloc = new.alloc[keep]
        new.names = [name for i, name in enumerate(new.names) if keep[i]]

    return repair_allocation(new, data, rng)


def op_split_pack(sol: Solution, data: ProblemData, rng: random.Random) -> Solution:
    if sol.content.shape[0] == 0:
        return sol

    new = copy.deepcopy(sol)
    candidates = np.flatnonzero(new.content.sum(axis=1) >= 2)
    if len(candidates) == 0:
        return sol

    p = int(rng.choice(list(candidates)))
    pack = new.content[p].copy()
    nonzero = np.flatnonzero(pack > 0)

    if len(nonzero) <= 1:
        # Split quantity of a single SKU.
        s = int(nonzero[0])
        q1 = max(1, pack[s] // 2)
        q2 = pack[s] - q1
        if q2 <= 0:
            return sol
        pack1 = np.zeros_like(pack)
        pack2 = np.zeros_like(pack)
        pack1[s] = q1
        pack2[s] = q2
    else:
        rng.shuffle(nonzero)
        cut = rng.randint(1, len(nonzero) - 1)
        idx1 = nonzero[:cut]
        idx2 = nonzero[cut:]
        pack1 = np.zeros_like(pack)
        pack2 = np.zeros_like(pack)
        pack1[idx1] = pack[idx1]
        pack2[idx2] = pack[idx2]

    # Replace original by two split packs.
    keep = np.ones(new.content.shape[0], dtype=bool)
    keep[p] = False

    old_alloc = new.alloc[p].copy()

    new.content = new.content[keep]
    new.alloc = new.alloc[keep]
    new.names = [name for i, name in enumerate(new.names) if keep[i]]

    append_pack(new, pack1, f"split_a_{p}")
    new.alloc[-1] = old_alloc.copy()

    append_pack(new, pack2, f"split_b_{p}")
    new.alloc[-1] = old_alloc.copy()

    return repair_allocation(new, data, rng)


def op_reallocate(sol: Solution, data: ProblemData, rng: random.Random) -> Solution:
    if sol.content.shape[0] == 0:
        return sol

    new = copy.deepcopy(sol)

    for _ in range(rng.randint(1, 10)):
        p = rng.randrange(new.content.shape[0])
        c = rng.randrange(new.alloc.shape[1])

        if rng.random() < 0.50 and new.alloc[p, c] > 0:
            new.alloc[p, c] -= rng.randint(1, min(3, new.alloc[p, c]))
        else:
            new.alloc[p, c] += rng.randint(1, 3)

    return repair_allocation(new, data, rng)


OPERATORS = [
    ("add_curve_pack", op_add_curve_pack, 1.30),
    ("remove_pack", op_remove_pack, 1.00),
    ("mutate_pack_content", op_mutate_pack_content, 1.40),
    ("merge_packs", op_merge_packs, 1.10),
    ("split_pack", op_split_pack, 0.70),
    ("reallocate", op_reallocate, 0.90),
]


def choose_operator(rng: random.Random, weights: Dict[str, float]):
    names = [name for name, _, _ in OPERATORS]
    funcs = {name: func for name, func, _ in OPERATORS}
    probs = [weights[name] for name in names]
    name = rng.choices(names, weights=probs, k=1)[0]
    return name, funcs[name]


# =============================================================================
# Metaheuristic optimizer
# =============================================================================

def optimize_metaheuristic(
    data: ProblemData,
    iterations: int,
    restarts: int,
    seed: int,
    initial_temperature: float,
    cooling_rate: float,
    n_seed_packs: int,
    time_limit: Optional[int],
    max_pack_types: Optional[int],
) -> Tuple[Solution, pd.DataFrame]:
    start_time = time.time()
    global_rng = random.Random(seed)

    best_global: Optional[Solution] = None
    log_rows = []

    base_operator_weights = {name: weight for name, _, weight in OPERATORS}

    for restart in range(1, restarts + 1):
        rng = random.Random(global_rng.randint(1, 10**9))

        current = greedy_construct_solution(data, rng, n_seed_packs=n_seed_packs)
        current = enforce_pack_type_limit(current, data, rng, max_pack_types)
        current = repair_allocation(current, data, rng)
        evaluate(current, data)

        best_restart = copy.deepcopy(current)
        temperature = initial_temperature

        operator_weights = base_operator_weights.copy()

        for it in range(1, iterations + 1):
            if time_limit is not None and time.time() - start_time >= time_limit:
                break

            op_name, op_func = choose_operator(rng, operator_weights)

            candidate = op_func(current, data, rng)
            candidate = enforce_pack_type_limit(candidate, data, rng, max_pack_types)
            candidate = repair_allocation(candidate, data, rng)
            evaluate(candidate, data)

            delta = candidate.cost - current.cost

            accepted = False
            if delta <= 0:
                accepted = True
            else:
                # Simulated annealing acceptance.
                prob = math.exp(-delta / max(temperature, 1e-9))
                accepted = rng.random() < prob

            if accepted:
                current = candidate

            if candidate.cost < best_restart.cost:
                best_restart = copy.deepcopy(candidate)

                # Reward successful operator.
                operator_weights[op_name] = operator_weights.get(op_name, 1.0) * 1.03

            if best_global is None or best_restart.cost < best_global.cost:
                best_global = copy.deepcopy(best_restart)

            if it % 250 == 0 or it == 1:
                log_rows.append({
                    "restart": restart,
                    "iteration": it,
                    "operator": op_name,
                    "current_cost": current.cost,
                    "restart_best_cost": best_restart.cost,
                    "global_best_cost": best_global.cost if best_global else np.nan,
                    "temperature": temperature,
                    "pack_types_current": int(current.content.shape[0]),
                    "pack_types_best": int(best_restart.content.shape[0]),
                    "shortage_units_best": int(best_restart.shortage_units),
                    "overstock_units_best": int(best_restart.overstock_units),
                    "elapsed_seconds": time.time() - start_time,
                })

            temperature *= cooling_rate

        print(
            f"Restart {restart:>2}/{restarts}: "
            f"best cost={best_restart.cost:,.2f}, "
            f"packs={best_restart.content.shape[0]}, "
            f"allocated={best_restart.allocated_packs:,}, "
            f"over={best_restart.overstock_units:,}, "
            f"short={best_restart.shortage_units:,}"
        )

        if time_limit is not None and time.time() - start_time >= time_limit:
            print("Time limit reached.")
            break

    if best_global is None:
        raise RuntimeError("Metaheuristic did not produce a solution.")

    best_global = repair_allocation(best_global, data, random.Random(seed + 999))
    evaluate(best_global, data)

    run_log = pd.DataFrame(log_rows)
    return best_global, run_log


def enforce_pack_type_limit(
    sol: Solution,
    data: ProblemData,
    rng: random.Random,
    max_pack_types: Optional[int],
) -> Solution:
    if max_pack_types is None:
        return sol

    sol = remove_empty_pack_rows(sol)

    while sol.content.shape[0] > max_pack_types:
        evaluate(sol, data)
        usage = sol.alloc.sum(axis=1)

        # Remove the least used pack type. Ties are randomized.
        min_usage = usage.min()
        candidates = np.flatnonzero(usage == min_usage)
        idx = int(rng.choice(list(candidates)))

        keep = np.ones(sol.content.shape[0], dtype=bool)
        keep[idx] = False

        sol.content = sol.content[keep]
        sol.alloc = sol.alloc[keep]
        sol.names = [name for i, name in enumerate(sol.names) if keep[i]]

        sol = repair_allocation(sol, data, rng)

    return sol


# =============================================================================
# Export
# =============================================================================

def solution_to_frames(sol: Solution, data: ProblemData) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    evaluate(sol, data)
    shipped = shipped_matrix(sol, data)

    active = (sol.content.sum(axis=1) > 0) & (sol.alloc.sum(axis=1) > 0)
    content = sol.content[active]
    alloc = sol.alloc[active]
    names = [name for name, keep in zip(sol.names, active) if keep]

    pack_ids = [f"P{i + 1:03d}" for i in range(content.shape[0])]

    packs_rows = []
    for pack_id, name, row in zip(pack_ids, names, content):
        rec = {"pack_id": pack_id, "pack_name": name}
        for sku, qty in zip(data.sku_ids, row):
            rec[sku] = int(qty)
        packs_rows.append(rec)
    packs_df = pd.DataFrame(packs_rows)

    alloc_rows = []
    for pack_id, row in zip(pack_ids, alloc):
        rec = {"pack_id": pack_id}
        for channel, qty in zip(data.channel_ids, row):
            rec[channel] = int(qty)
        alloc_rows.append(rec)
    allocation_df = pd.DataFrame(alloc_rows)

    assortment_df = pd.DataFrame(shipped, index=data.sku_ids, columns=data.channel_ids)
    assortment_df.insert(0, "sku_id", data.sku_ids)

    diag_rows = []
    for s_idx, sku in enumerate(data.sku_ids):
        for c_idx, channel in enumerate(data.channel_ids):
            forecast = int(data.demand[s_idx, c_idx])
            sent = int(shipped[s_idx, c_idx])
            over = max(0, sent - forecast)
            under = max(0, forecast - sent)
            diag_rows.append({
                "sku_id": sku,
                "channel_id": channel,
                "forecast_units": forecast,
                "shipped_units": sent,
                "overstock_units": over,
                "shortage_units": under,
                "unit_cost": float(data.unit_cost[s_idx]),
                "capital_cost": over * float(data.unit_cost[s_idx]) * COST_OF_CAPITAL,
            })

    diagnostics_df = pd.DataFrame(diag_rows)

    total_forecast = int(data.demand.sum())
    total_shipped = int(shipped.sum())
    total_over = int(diagnostics_df["overstock_units"].sum())
    total_under = int(diagnostics_df["shortage_units"].sum())
    total_allocated_packs = int(alloc.sum())
    total_pack_types = int(content.shape[0])

    setup_cost = total_pack_types * PACK_CREATION_COST
    handling_cost = total_allocated_packs * HANDLING_COST_PER_PACK
    capital_cost = float(diagnostics_df["capital_cost"].sum())
    shortage_cost = total_under * SHORTAGE_PENALTY_PER_UNIT
    total_cost = setup_cost + handling_cost + capital_cost + shortage_cost

    baseline_no_packs_cost = total_forecast * HANDLING_COST_PER_PACK

    cost_summary_df = pd.DataFrame([
        {"metric": "total_cost", "value": total_cost},
        {"metric": "setup_cost", "value": setup_cost},
        {"metric": "handling_cost", "value": handling_cost},
        {"metric": "capital_cost", "value": capital_cost},
        {"metric": "shortage_cost", "value": shortage_cost},
        {"metric": "pack_types_used", "value": total_pack_types},
        {"metric": "allocated_packs_total", "value": total_allocated_packs},
        {"metric": "forecast_units_total", "value": total_forecast},
        {"metric": "shipped_units_total", "value": total_shipped},
        {"metric": "overstock_units_total", "value": total_over},
        {"metric": "shortage_units_total", "value": total_under},
        {"metric": "baseline_no_packs_handling_cost", "value": baseline_no_packs_cost},
        {"metric": "estimated_savings_vs_unit_handling_only", "value": baseline_no_packs_cost - total_cost},
    ])

    return packs_df, allocation_df, assortment_df, diagnostics_df, cost_summary_df


def write_clean_solution_workbook(
    output_path: Path,
    packs_df: pd.DataFrame,
    allocation_df: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Keep only the two solution tabs if no official template is supplied.
        packs_out = packs_df.drop(columns=["pack_name"], errors="ignore")
        packs_out.to_excel(writer, sheet_name="Packs", index=False)
        allocation_df.to_excel(writer, sheet_name="PackAllocation", index=False)


def fill_template_workbook(
    template_path: Path,
    output_path: Path,
    packs_df: pd.DataFrame,
    allocation_df: pd.DataFrame,
) -> bool:
    """
    Best-effort template filling. If the template structure cannot be detected,
    return False and let the caller create a clean workbook.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return False

    if not template_path.exists():
        return False

    wb = load_workbook(template_path)

    if "Packs" not in wb.sheetnames or "PackAllocation" not in wb.sheetnames:
        return False

    def detect_header(ws, expected_values: Sequence[str], max_rows: int = 25):
        expected_set = set(map(str, expected_values))
        best_row = None
        best_hits = 0
        for r in range(1, min(ws.max_row, max_rows) + 1):
            vals = [str(cell.value) for cell in ws[r] if cell.value not in (None, "")]
            hits = len(set(vals) & expected_set)
            if hits > best_hits:
                best_hits = hits
                best_row = r
        return best_row, best_hits

    ws_packs = wb["Packs"]
    sku_cols = [c for c in packs_df.columns if c not in {"pack_id", "pack_name"}]
    header_row, hits = detect_header(ws_packs, sku_cols)

    if header_row is None or hits == 0:
        return False

    sku_col_map = {}
    for cell in ws_packs[header_row]:
        value = str(cell.value) if cell.value not in (None, "") else None
        if value in sku_cols:
            sku_col_map[value] = cell.column

    pack_start = header_row + 1

    for i, (_, row) in enumerate(packs_df.iterrows()):
        excel_row = pack_start + i
        ws_packs.cell(excel_row, 1).value = row["pack_id"]
        for sku, col in sku_col_map.items():
            ws_packs.cell(excel_row, col).value = int(row.get(sku, 0))

    # Zero unused template pack rows.
    for excel_row in range(pack_start + len(packs_df), ws_packs.max_row + 1):
        for col in sku_col_map.values():
            ws_packs.cell(excel_row, col).value = 0

    ws_alloc = wb["PackAllocation"]
    channels = [c for c in allocation_df.columns if c != "pack_id"]
    header_row_a, hits_a = detect_header(ws_alloc, channels)

    if header_row_a is None or hits_a == 0:
        return False

    ch_col_map = {}
    for cell in ws_alloc[header_row_a]:
        value = str(cell.value) if cell.value not in (None, "") else None
        if value in channels:
            ch_col_map[value] = cell.column

    alloc_start = header_row_a + 1
    alloc_lookup = allocation_df.set_index("pack_id")

    for i, (_, pack_row) in enumerate(packs_df.iterrows()):
        pack_id = pack_row["pack_id"]
        excel_row = alloc_start + i
        ws_alloc.cell(excel_row, 1).value = pack_id
        for ch, col in ch_col_map.items():
            value = 0
            if pack_id in alloc_lookup.index:
                value = int(alloc_lookup.loc[pack_id].get(ch, 0))
            ws_alloc.cell(excel_row, col).value = value

    for excel_row in range(alloc_start + len(packs_df), ws_alloc.max_row + 1):
        for col in ch_col_map.values():
            ws_alloc.cell(excel_row, col).value = 0

    wb.save(output_path)
    return True


def export_solution(
    sol: Solution,
    data: ProblemData,
    output_dir: str,
    template_path: str,
    run_log: pd.DataFrame,
) -> None:
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    packs_df, allocation_df, assortment_df, diagnostics_df, cost_summary_df = solution_to_frames(sol, data)

    packs_df.to_csv(out / "optimalisation_packs_metaheuristic.csv", index=False)
    allocation_df.to_csv(out / "optimalisation_pack_allocation_metaheuristic.csv", index=False)
    assortment_df.to_csv(out / "optimalisation_assortment_metaheuristic.csv", index=False)
    diagnostics_df.to_csv(out / "optimalisation_diagnostics_metaheuristic.csv", index=False)
    cost_summary_df.to_csv(out / "optimalisation_cost_summary_metaheuristic.csv", index=False)

    if not run_log.empty:
        run_log.to_csv(out / "optimalisation_run_log.csv", index=False)

    workbook_path = out / "optimalisation_solution.xlsx"
    template_file = Path(template_path)

    filled = fill_template_workbook(
        template_path=template_file,
        output_path=workbook_path,
        packs_df=packs_df,
        allocation_df=allocation_df,
    )

    if filled:
        print(f"Filled official template: {workbook_path}")
    else:
        write_clean_solution_workbook(workbook_path, packs_df, allocation_df)
        print(f"Created clean solution workbook: {workbook_path}")

    print("\nCost summary:")
    print(cost_summary_df.to_string(index=False))


# =============================================================================
# CLI
# =============================================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Metaheuristic pre-pack content and allocation optimization."
    )

    parser.add_argument(
        "--forecast",
        default="outputs/master_forecast_2026.csv",
        help="Forecast CSV generated by run_forecasting.py.",
    )
    parser.add_argument(
        "--products",
        default="data/PPP_stu_products.xlsx",
        help="Product master file with product costs.",
    )
    parser.add_argument(
        "--template",
        default="data/PPP_solutionFile_2026.xlsx",
        help="Official solution template, if available.",
    )
    parser.add_argument(
        "--output-dir",
        default="outputs",
        help="Output directory.",
    )
    parser.add_argument(
        "--iterations",
        type=int,
        default=20000,
        help="Iterations per restart.",
    )
    parser.add_argument(
        "--restarts",
        type=int,
        default=5,
        help="Number of independent restarts.",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=42,
        help="Random seed.",
    )
    parser.add_argument(
        "--initial-temperature",
        type=float,
        default=5000.0,
        help="Initial simulated annealing temperature.",
    )
    parser.add_argument(
        "--cooling-rate",
        type=float,
        default=0.9995,
        help="Temperature multiplier per iteration.",
    )
    parser.add_argument(
        "--seed-packs",
        type=int,
        default=250,
        help="Number of demand-curve packs added during initial construction.",
    )
    parser.add_argument(
        "--time-limit",
        type=int,
        default=None,
        help="Optional total runtime limit in seconds.",
    )
    parser.add_argument(
        "--max-pack-types",
        type=int,
        default=MAX_PACK_TYPES_DEFAULT,
        help="Optional maximum number of pack types in the final solution.",
    )
    parser.add_argument(
        "--min-forecast",
        type=int,
        default=1,
        help="Ignore SKUs with total forecast below this value.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()
    seed_everything(args.seed)

    data = load_problem(
        forecast_path=args.forecast,
        product_path=args.products,
        min_forecast=args.min_forecast,
    )

    print("=" * 78)
    print("METAHEURISTIC PRE-PACK OPTIMIZATION")
    print("=" * 78)
    print(f"Forecast file        : {args.forecast}")
    print(f"SKUs                 : {len(data.sku_ids):,}")
    print(f"Sales channels       : {len(data.channel_ids):,}")
    print(f"Forecast units total : {int(data.demand.sum()):,}")
    print(f"Max pack units       : {MAX_PACK_UNITS}")
    print(f"Max SKUs per pack    : {MAX_DISTINCT_SKUS_PER_PACK}")
    print(f"Max pack types       : {args.max_pack_types if args.max_pack_types is not None else 'unlimited'}")
    print(f"Iterations/restarts  : {args.iterations:,} x {args.restarts}")
    print("=" * 78)

    best, run_log = optimize_metaheuristic(
        data=data,
        iterations=args.iterations,
        restarts=args.restarts,
        seed=args.seed,
        initial_temperature=args.initial_temperature,
        cooling_rate=args.cooling_rate,
        n_seed_packs=args.seed_packs,
        time_limit=args.time_limit,
        max_pack_types=args.max_pack_types,
    )

    #export_solution(
    #    sol=best,
    #    data=data,
    #    output_dir=args.output_dir,
    #    template_path=args.template,
    #    run_log=run_log,
    #)

    print("\nFinal solution:")
    print(f"  total cost       : {best.cost:,.2f}")
    print(f"  setup cost       : {best.setup_cost:,.2f}")
    print(f"  handling cost    : {best.handling_cost:,.2f}")
    print(f"  capital cost     : {best.capital_cost:,.2f}")
    print(f"  shortage cost    : {best.shortage_cost:,.2f}")
    print(f"  pack types       : {best.content.shape[0]:,}")
    print(f"  allocated packs  : {best.allocated_packs:,}")
    print(f"  overstock units  : {best.overstock_units:,}")
    print(f"  shortage units   : {best.shortage_units:,}")
    print("\nDone.")


if __name__ == "__main__":
    main()
