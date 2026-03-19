# run_forecasting.py
# Runs the full forecasting pipeline for ALL product categories.
# For each category:
#   1. Lost Sales Correction (hierarchical, 3 levels)
#   2. MAPE-based ensemble weights (MA, SES, Holt, LinReg)
#   3. Forecast 2026 per channel
#   4. Size distribution via historical shares
#   5. Excel report + CSVs + plots
#
# Input:  outputs/merged_data.csv
#         data/PPP_stu_product_categories.xlsx
# Output: outputs/forecast_<category>_results.xlsx  (per category)
#         outputs/forecast_<category>_channel_total.csv
#         outputs/forecast_<category>_channel_size.csv
#         outputs/plots_<category>/

import os
import re
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.linear_model import LinearRegression
from statsmodels.tsa.holtwinters import ExponentialSmoothing, Holt
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from lost_sales_correction import apply_lost_sales_correction

warnings.filterwarnings('ignore')
os.makedirs('outputs', exist_ok=True)

VALIDATION_YEAR = 2025
SES_ALPHA       = 0.47
MA_WINDOW       = 3

SIZE_ORDER_WOMEN = ['XXS', 'XS', 'S', 'M', 'L', 'XL']
SIZE_ORDER_MEN   = ['XS',  'S',  'M', 'L', 'XL', 'XXL']
SIZE_ORDER_ONE   = ['onesize']

HEADER_COLOR = '1F4E79'
TOTAL_COLOR  = 'D6E4F0'
TITLE_COLOR  = '2E75B6'

# ── Load categories from official product file ────────────────────────────────

def load_categories(path='data/PPP_stu_product_categories.xlsx') -> dict:
    prod = pd.read_excel(path, header=0)
    prod.columns = ['id', 'name', 'category_group', 'category', 'price', 'cost']
    prod['category_group'] = prod['category_group'].ffill()
    return {grp: sub['id'].tolist()
            for grp, sub in prod.groupby('category_group')}


def slugify(name: str) -> str:
    """'T-shirt men' → 'tshirt_men'"""
    return re.sub(r'[^a-z0-9]+', '_', name.lower()).strip('_')


def get_size_order(df_cat: pd.DataFrame) -> list:
    sizes = df_cat['size'].unique()
    if 'onesize' in sizes:
        return SIZE_ORDER_ONE
    elif 'XXL' in sizes:
        return [s for s in SIZE_ORDER_MEN if s in sizes]
    else:
        return [s for s in SIZE_ORDER_WOMEN if s in sizes]


# ── Core forecasting function ─────────────────────────────────────────────────

def forecast_category(cat_name: str, product_ids: list, df_all: pd.DataFrame):

    slug  = slugify(cat_name)
    label = cat_name
    print(f"\n{'#'*65}")
    print(f"  {label}")
    print(f"{'#'*65}")

    os.makedirs(f'outputs/plots_{slug}', exist_ok=True)

    # ── 1. Filter & Lost Sales Correction ────────────────────────────────────
    df = df_all[df_all['product_id'].isin(product_ids)].copy()
    df = apply_lost_sales_correction(df, label=label)

    size_order = get_size_order(df)

    # ── 2. Aggregate to channel × season (corrected units) ───────────────────
    channel_season = (df.groupby(['channel_id', 'season'])['units_corrected']
                        .sum().reset_index()
                        .rename(columns={'units_corrected': 'total_units'}))

    channels = sorted(channel_season['channel_id'].unique())
    seasons  = sorted(channel_season['season'].unique())
    print(f"\n  {len(channels)} channels, seasons {min(seasons)}–{max(seasons)}")

    # ── 3. MAPE-based ensemble weights ───────────────────────────────────────
    mape_ma, mape_ses, mape_holt, mape_lr = [], [], [], []

    for channel in channels:
        ch     = channel_season[channel_season['channel_id'] == channel].sort_values('season')
        y, yr  = ch['total_units'].values.astype(float), ch['season'].values
        train  = y[yr < VALIDATION_YEAR]
        actual = y[yr == VALIDATION_YEAR]
        if len(actual) == 0 or actual[0] == 0 or len(train) < MA_WINDOW:
            continue
        actual = actual[0]
        yr_tr  = yr[yr < VALIDATION_YEAR]

        mape_ma.append(abs(np.mean(train[-MA_WINDOW:]) - actual) / actual * 100)
        ses = ExponentialSmoothing(train, trend=None, seasonal=None).fit(
                  smoothing_level=SES_ALPHA, optimized=False)
        mape_ses.append(abs(float(ses.forecast(1)[0]) - actual) / actual * 100)
        holt = Holt(train).fit()
        mape_holt.append(abs(float(holt.forecast(1)[0]) - actual) / actual * 100)
        lr = LinearRegression().fit(yr_tr.reshape(-1, 1), train)
        mape_lr.append(abs(float(lr.predict([[VALIDATION_YEAR]])[0]) - actual) / actual * 100)

    avg_mapes = np.array([np.mean(mape_ma), np.mean(mape_ses),
                          np.mean(mape_holt), np.mean(mape_lr)])
    raw_w   = 1.0 / avg_mapes
    weights = raw_w / raw_w.sum()

    METHOD_NAMES = [f'MA({MA_WINDOW})', f'SES (α={SES_ALPHA})',
                    "Holt's Linear Trend", 'Linear Regression']

    print(f"\n  Method validation (MAPE on {VALIDATION_YEAR}):")
    for name, mape, w in zip(METHOD_NAMES, avg_mapes, weights):
        print(f"    {name:<22}  MAPE = {mape:5.1f}%   weight = {w*100:.1f}%")

    # ── 4. Forecast 2026 per channel ─────────────────────────────────────────
    channel_results = []

    for channel in channels:
        ch     = channel_season[channel_season['channel_id'] == channel].sort_values('season')
        y, yr  = ch['total_units'].values.astype(float), ch['season'].values

        ma_val   = float(np.mean(y[-MA_WINDOW:]))
        ses_val  = float(ExponentialSmoothing(y, trend=None, seasonal=None)
                         .fit(smoothing_level=SES_ALPHA, optimized=False).forecast(1)[0])
        holt_val = float(Holt(y).fit().forecast(1)[0])
        lr_val   = float(LinearRegression()
                         .fit(yr.reshape(-1, 1), y).predict([[2026]])[0])

        ensemble = float(np.dot(weights, [ma_val, ses_val, holt_val, lr_val]))
        ensemble = max(0.0, ensemble)

        channel_results.append({
            'channel_id':        channel,
            'forecast_MA':       round(max(0, ma_val)),
            'forecast_SES':      round(max(0, ses_val)),
            'forecast_Holt':     round(max(0, holt_val)),
            'forecast_LinReg':   round(max(0, lr_val)),
            'forecast_ensemble': round(ensemble),
        })

        # Plot
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.plot(yr, y, marker='o', color='black', lw=2, label='Observed')
        ax.plot([yr[-1], 2026], [y[-1], ma_val],   marker='s', color='#2a6db5',
                lw=1.5, label=f'MA ({round(ma_val)})')
        ax.plot([yr[-1], 2026], [y[-1], ses_val],  marker='^', color='#e67e22',
                lw=1.5, label=f'SES ({round(ses_val)})')
        ax.plot([yr[-1], 2026], [y[-1], holt_val], marker='D', color='#27ae60',
                lw=1.5, label=f"Holt ({round(holt_val)})")
        ax.plot([yr[-1], 2026], [y[-1], lr_val],   marker='v', color='#8e44ad',
                lw=1.5, label=f'LinReg ({round(lr_val)})')
        ax.scatter(2026, ensemble, color='red', s=120, zorder=5,
                   label=f'Ensemble ({round(ensemble)})')
        ax.set_title(f'{label} – {channel} – Forecast 2026',
                     fontsize=13, fontweight='bold')
        ax.set_xlabel('Season')
        ax.set_ylabel(f'Units ({len(product_ids)} products, all sizes)')
        ax.set_xticks(list(range(int(yr.min()), 2027)))
        ax.legend(fontsize=8, bbox_to_anchor=(1.01, 1), loc='upper left')
        ax.grid(True, linestyle=':', alpha=0.6)
        plt.tight_layout()
        plt.savefig(f'outputs/plots_{slug}/forecast_{channel}.png', dpi=150)
        plt.close()

    df_channel = pd.DataFrame(channel_results)
    print(f"\n  Total ensemble 2026: {df_channel['forecast_ensemble'].sum():,} units")

    # ── 5. Size distribution ──────────────────────────────────────────────────
    size_shares = (df.groupby(['channel_id', 'size'])['units_corrected']
                     .sum().reset_index()
                     .rename(columns={'units_corrected': 'hist_units'}))
    ch_totals = size_shares.groupby('channel_id')['hist_units'].transform('sum')
    size_shares['share'] = size_shares['hist_units'] / ch_totals

    available_sizes = [s for s in size_order if s in size_shares['size'].unique()]

    rows = []
    for _, ch_row in df_channel.iterrows():
        channel  = ch_row['channel_id']
        total_fc = ch_row['forecast_ensemble']
        ch_sizes = size_shares[size_shares['channel_id'] == channel].set_index('size')['share']
        for size in available_sizes:
            rows.append({
                'channel_id':        channel,
                'size':              size,
                'forecast_ensemble': round(total_fc * ch_sizes.get(size, 0.0)),
                'size_share':        round(ch_sizes.get(size, 0.0), 4),
            })

    df_channel_size = pd.DataFrame(rows)

    # ── 6. Export CSVs ────────────────────────────────────────────────────────
    df_channel.to_csv(f'outputs/forecast_{slug}_channel_total.csv', index=False)
    df_channel_size.to_csv(f'outputs/forecast_{slug}_channel_size.csv', index=False)

    # ── 7. Size-level sanity check ────────────────────────────────────────────
    size_season = (df.groupby(['size', 'season'])['units_corrected']
                     .sum().reset_index()
                     .rename(columns={'units_corrected': 'total_units'}))

    size_fc_rows = []
    for size in available_sizes:
        sz     = size_season[size_season['size'] == size].sort_values('season')
        y, yr  = sz['total_units'].values.astype(float), sz['season'].values
        ma_v   = float(np.mean(y[-MA_WINDOW:]))
        ses_v  = float(ExponentialSmoothing(y, trend=None, seasonal=None)
                       .fit(smoothing_level=SES_ALPHA, optimized=False).forecast(1)[0])
        holt_v = float(Holt(y).fit().forecast(1)[0])
        lr_v   = float(LinearRegression()
                       .fit(yr.reshape(-1, 1), y).predict([[2026]])[0])
        ens_v  = float(np.dot(weights, [ma_v, ses_v, holt_v, lr_v]))
        size_fc_rows.append({
            'Size':              size,
            f'MA({MA_WINDOW})':  round(max(0, ma_v)),
            f'SES (α={SES_ALPHA})': round(max(0, ses_v)),
            "Holt's":            round(max(0, holt_v)),
            'Linear Regression': round(max(0, lr_v)),
            'Ensemble':          round(max(0, ens_v)),
        })

    df_size_fc = pd.DataFrame(size_fc_rows)
    totals = df_size_fc.drop(columns='Size').sum()
    totals['Size'] = 'TOTAL'
    df_size_fc = pd.concat([df_size_fc, pd.DataFrame([totals])], ignore_index=True)

    # ── 8. Excel export ───────────────────────────────────────────────────────
    df_weights = pd.DataFrame({
        'Method':              METHOD_NAMES,
        'Avg MAPE 2025 (%)':   [round(m, 1) for m in avg_mapes],
        'Inverse-MAPE Weight': [round(w, 3) for w in weights],
        'Weight (%)':          [round(w * 100, 1) for w in weights],
        'Description': [
            f'Simple average of last {MA_WINDOW} observed years',
            f'Exponential smoothing, more weight on recent years (α={SES_ALPHA})',
            'Accounts for linear trend, parameters auto-optimized',
            'OLS linear regression over all 8 years (2018–2025)',
        ],
    })

    hist_pivot = channel_season.pivot(
        index='channel_id', columns='season', values='total_units')
    hist_pivot.columns = [str(c) for c in hist_pivot.columns]

    df_ch_idx = df_channel.set_index('channel_id').rename(columns={
        'forecast_MA':       f'MA({MA_WINDOW})',
        'forecast_SES':      f'SES (α={SES_ALPHA})',
        'forecast_Holt':     "Holt's",
        'forecast_LinReg':   'LinReg',
        'forecast_ensemble': 'Ensemble 2026',
    })

    per_channel_tbl = pd.concat([hist_pivot, df_ch_idx], axis=1).round(0).astype(int)
    total_row = per_channel_tbl.sum().rename('TOTAL')
    per_channel_tbl = pd.concat([per_channel_tbl, total_row.to_frame().T])
    per_channel_tbl.index.name = 'Channel'
    per_channel_tbl = per_channel_tbl.reset_index()

    pivot_fc = df_channel_size.pivot(
        index='channel_id', columns='size', values='forecast_ensemble')[available_sizes]
    pivot_fc['Total'] = pivot_fc.sum(axis=1)
    ch_sz_tbl = pivot_fc.copy()
    total_row_sz = ch_sz_tbl.sum().rename('TOTAL')
    ch_sz_tbl = pd.concat([ch_sz_tbl, total_row_sz.to_frame().T])
    ch_sz_tbl.index.name = 'Channel'
    ch_sz_tbl = ch_sz_tbl.reset_index()

    excel_out = f'outputs/forecast_{slug}_results.xlsx'

    with pd.ExcelWriter(excel_out, engine='openpyxl') as writer:
        df_weights.to_excel(     writer, sheet_name='Method Weights',    index=False, startrow=3)
        per_channel_tbl.to_excel(writer, sheet_name='Per Channel',       index=False, startrow=3)
        ch_sz_tbl.to_excel(      writer, sheet_name='Channel x Size',    index=False, startrow=3)
        df_size_fc.to_excel(     writer, sheet_name='Size Verification', index=False, startrow=3)

        sheets_meta = {
            'Method Weights':    (f'{label} – Forecast Method Evaluation (Validation Year: {VALIDATION_YEAR})',
                                  f'Train: 2018–{VALIDATION_YEAR-1}  |  Validate: {VALIDATION_YEAR}  |  MAPE averaged over channels  →  inverse-MAPE weights'),
            'Per Channel':       (f'{label} – Historical Demand per Channel + Forecast 2026',
                                  f'Historical: units sold per channel (2018–{VALIDATION_YEAR})  |  Forecast 2026: 4 methods + MAPE-weighted ensemble'),
            'Channel x Size':    (f'{label} – Forecast 2026 per Channel × Size  [Main Output for Optimization]',
                                  'Channel totals (Ensemble) distributed via historical size-share weights per channel'),
            'Size Verification': (f'{label} – Size-Level Forecast 2026 (Sanity Check)',
                                  'Aggregated across all channels, forecasted per size  |  Uses same MAPE weights as main model'),
        }

        header_font  = Font(bold=True, color='FFFFFF')
        header_fill  = PatternFill('solid', fgColor=HEADER_COLOR)
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        total_fill   = PatternFill('solid', fgColor=TOTAL_COLOR)
        total_font   = Font(bold=True)

        has_total_row = {'Per Channel', 'Channel x Size', 'Size Verification'}

        for sheet_name, (title, subtitle) in sheets_meta.items():
            ws = writer.sheets[sheet_name]
            ws['A1'] = title
            ws['A1'].font = Font(bold=True, size=13, color=TITLE_COLOR)
            ws['A2'] = subtitle
            ws['A2'].font = Font(italic=True, size=10, color='595959')
            for cell in ws[4]:
                if cell.value is not None:
                    cell.font      = header_font
                    cell.fill      = header_fill
                    cell.alignment = header_align
            if sheet_name in has_total_row:
                last_row = ws.max_row
                for cell in ws[last_row]:
                    cell.font = total_font
                    cell.fill = total_fill
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
                ws.column_dimensions[col_letter].width = min(max_len + 3, 25)
            ws.freeze_panes = 'A5'

    print(f"  Saved: {excel_out}")


# ── Main: run all categories ──────────────────────────────────────────────────

if __name__ == '__main__':
    df_all     = pd.read_csv('outputs/merged_data.csv')
    categories = load_categories()

    print(f"Found {len(categories)} categories: {list(categories.keys())}")

    summary = []
    for cat_name, product_ids in categories.items():
        forecast_category(cat_name, product_ids, df_all)
        slug = slugify(cat_name)
        try:
            total = pd.read_csv(f'outputs/forecast_{slug}_channel_total.csv')
            summary.append({
                'Category':      cat_name,
                'Products':      len(product_ids),
                'Forecast 2026': total['forecast_ensemble'].sum(),
            })
        except FileNotFoundError:
            pass

    print(f"\n{'='*50}")
    print("  SUMMARY — Forecast 2026 all categories")
    print(f"{'='*50}")
    df_sum = pd.DataFrame(summary)
    print(df_sum.to_string(index=False))
    print(f"\n  Grand total: {df_sum['Forecast 2026'].sum():,} units")
    print(f"{'='*50}")
