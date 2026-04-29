#!/usr/bin/env python3
"""
optimalisation.py

Pre-pack optimization for the Sustainable Essentials SCM group assignment.

This script takes the 2026 SKU/channel forecast created by run_forecasting.py and
builds an integer optimization model for the number and content of pre-packs and
their allocation to each sales channel.

Required solver:
    Gurobi, via gurobipy

Main inputs expected:
    outputs/master_forecast_2026.csv
        Required columns:
            channel_id, product_id, size, forecast_ensemble
        Optional columns:
            category_group

    data/PPP_stu_products.xlsx
        Used to obtain SKU unit costs for capital-cost calculations.
        If unavailable, the script falls back to unit cost = 1 for all SKUs.

Optional input:
    data/PPP_solutionFile_2026.xlsx
        If present, the script fills the existing template while preserving its
        workbook structure. If not present, it creates a clean solution workbook.

Main outputs:
    outputs/optimalisation_solution.xlsx
    outputs/optimalisation_packs.csv
    outputs/optimalisation_pack_allocation.csv
    outputs/optimalisation_assortment.csv
    outputs/optimalisation_cost_summary.csv

Model idea:
    A pack type p has a fixed integer content vector a[p, sku].
    x[p, c] = integer number of packs p allocated to channel c.
    y[p]    = 1 if pack type p is used.

    shipped[sku, c] = sum_p a[p, sku] * x[p, c]

    The script generates a large set of candidate pack types first, then lets
    Gurobi choose the best subset and allocation.

    This avoids a nonlinear model because directly optimizing both pack content
    and pack allocation would create products of decision variables.

Author note:
    The filename is kept as "optimalisation.py" because that is what was requested.
"""

from __future__ import annotations

import argparse
import math
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import numpy as np
import pandas as pd

try:
    import gurobipy as gp
    from gurobipy import GRB
except ImportError as exc:
    raise ImportError(
        "This script requires Gurobi. Install gurobipy and make sure your "
        "Gurobi license is active before running this script."
    ) from exc


# ---------------------------------------------------------------------------
# Cost and model parameters
# ---------------------------------------------------------------------------

COST_OF_CAPITAL = 0.243
HANDLING_COST_PER_PACK = 11.03
PACK_CREATION_COST = 134.00

# Shortage is allowed by default with a very high penalty. This makes the model
# robust if the candidate pack set cannot cover every forecast exactly.
# Set --no-shortage to force all forecasted demand to be covered.
SHORTAGE_PENALTY_PER_UNIT = 10_000.00

# Candidate pack-generation limits.
MAX_PACK_UNITS = 10000
MAX_DISTINCT_SKUS_PER_PACK = 10000
MAX_CANDIDATE_PACKS = 50_000

# The template shown in the assignment starts with one pack row per SKU.
# Keeping at most one pack type per SKU is a good practical limit for the
# submitted solution file, but you can override it with --max-pack-types.
USE_AT_MOST_ONE_PACK_PER_SKU_BY_DEFAULT = False

# If True, the model will not allow any shortage of the forecasted demand.
# If False, the model will allow shortage with a penalty.
NO_SHORTAGE_BY_DEFAULT = True

# The default gap between the best-found solution and the linear-programming 
# relaxation bound before Gurobi stops early. Set to 0.1 (10%) to speed up testing; 
# use 0.01 (1%) or 0.005 (0.5%) or 0 for the final runs.
MIP_GAP = 0.01

TIME_LIMIT = 28000

# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class PackCandidate:
    """Sparse representation of one candidate pack type."""
    name: str
    content: Tuple[Tuple[str, int], ...]

    @property
    def total_units(self) -> int:
        return sum(qty for _, qty in self.content)

    @property
    def distinct_skus(self) -> int:
        return len(self.content)

    def as_dict(self) -> Dict[str, int]:
        return dict(self.content)


# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------

def slugify(value: str) -> str:
    """Create a safe identifier for pack names."""
    value = str(value).lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_")


def make_sku_id(product_id, size) -> str:
    """SKU identifier used in the solution workbook."""
    return f"{product_id}_{size}"


def ceil_int(x: float) -> int:
    """Round a forecast up to the nearest non-negative integer."""
    if pd.isna(x):
        return 0
    return max(0, int(math.ceil(float(x))))


def normalize_pack_content(raw: Dict[str, int]) -> Tuple[Tuple[str, int], ...]:
    """
    Convert a raw pack dictionary to a sorted immutable sparse representation.
    Removes zero entries.
    """
    clean = {str(k): int(v) for k, v in raw.items() if int(v) > 0}
    return tuple(sorted(clean.items()))


def add_candidate(
    candidates: Dict[Tuple[Tuple[str, int], ...], PackCandidate],
    name: str,
    raw_content: Dict[str, int],
) -> None:
    """Add a candidate pack if it is feasible and not already present."""
    content = normalize_pack_content(raw_content)
    if not content:
        return

    total_units = sum(qty for _, qty in content)
    distinct = len(content)

    if total_units > MAX_PACK_UNITS:
        return
    if distinct > MAX_DISTINCT_SKUS_PER_PACK:
        return

    if content not in candidates:
        candidates[content] = PackCandidate(name=name, content=content)


def proportional_integer_pack(
    demand: pd.Series,
    target_units: int,
    max_distinct_skus: int = MAX_DISTINCT_SKUS_PER_PACK,
) -> Dict[str, int]:
    """
    Create an integer pack following the demand proportions in a Series indexed
    by sku_id.

    The method:
        1. Keep the highest-demand SKUs.
        2. Allocate target_units proportionally.
        3. Use largest remainders to repair rounding.
    """
    demand = demand[demand > 0].sort_values(ascending=False).head(max_distinct_skus)
    if demand.empty or target_units <= 0:
        return {}

    shares = demand / demand.sum()
    raw = shares * target_units
    base = np.floor(raw).astype(int)

    # Ensure every selected SKU gets at least 1 unit if possible.
    selected = list(demand.index)
    if target_units >= len(selected):
        base[:] = np.maximum(base, 1)

    diff = target_units - int(base.sum())

    if diff > 0:
        remainders = (raw - np.floor(raw)).sort_values(ascending=False)
        for sku in list(remainders.index)[:diff]:
            base.loc[sku] += 1
    elif diff < 0:
        removable = base[base > 1].sort_values(ascending=False)
        for sku in list(removable.index):
            if diff == 0:
                break
            base.loc[sku] -= 1
            diff += 1

    return {sku: int(qty) for sku, qty in base.items() if int(qty) > 0}


# ---------------------------------------------------------------------------
# Load data
# ---------------------------------------------------------------------------

def load_forecast(path: str) -> pd.DataFrame:
    """Load the SKU/channel forecast and standardize required columns."""
    df = pd.read_csv(path)

    required = {"channel_id", "product_id", "size", "forecast_ensemble"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Forecast file is missing required columns: {sorted(missing)}")

    df = df.copy()
    df["sku_id"] = [make_sku_id(pid, size) for pid, size in zip(df["product_id"], df["size"])]
    df["channel_id"] = df["channel_id"].astype(str)
    df["forecast_units"] = df["forecast_ensemble"].apply(ceil_int)

    # Combine duplicates just in case several category files contributed rows.
    group_cols = ["channel_id", "sku_id", "product_id", "size"]
    if "category_group" in df.columns:
        group_cols.append("category_group")

    df = (
        df.groupby(group_cols, dropna=False)["forecast_units"]
          .sum()
          .reset_index()
    )

    return df


def load_unit_costs(products_path: str, forecast: pd.DataFrame) -> Dict[str, float]:
    """
    Load SKU unit costs. If the products file is not available, use cost = 1.
    The objective remains structurally correct, but cost magnitudes become less
    meaningful without real costs.
    """
    if not Path(products_path).exists():
        print(f"WARNING: {products_path} not found. Using unit cost = 1 for all SKUs.")
        return {sku: 1.0 for sku in forecast["sku_id"].unique()}

    products = pd.read_excel(products_path)
    if not {"id", "cost"}.issubset(products.columns):
        print("WARNING: products file has no columns 'id' and 'cost'. Using unit cost = 1.")
        return {sku: 1.0 for sku in forecast["sku_id"].unique()}

    cost_by_pid = products.set_index("id")["cost"].to_dict()

    costs = {}
    for _, row in forecast[["sku_id", "product_id"]].drop_duplicates().iterrows():
        costs[row["sku_id"]] = float(cost_by_pid.get(row["product_id"], 1.0))

    return costs


def build_demand_table(forecast: pd.DataFrame) -> pd.DataFrame:
    """Create a SKU x channel demand matrix."""
    demand = (
        forecast.pivot_table(
            index="sku_id",
            columns="channel_id",
            values="forecast_units",
            aggfunc="sum",
            fill_value=0,
        )
        .astype(int)
        .sort_index()
    )

    # Remove SKUs with zero total forecast, unless you want to force all template SKUs.
    demand = demand[demand.sum(axis=1) > 0]
    return demand


# ---------------------------------------------------------------------------
# Candidate pack generation
# ---------------------------------------------------------------------------

def generate_candidate_packs(
    forecast: pd.DataFrame,
    demand: pd.DataFrame,
    max_candidate_packs: int = MAX_CANDIDATE_PACKS,
) -> List[PackCandidate]:
    """
    Generate useful candidate pack types.

    Candidate families:
        1. Single-SKU packs with several quantities.
        2. Same-product size-curve packs based on total demand.
        3. Same-product channel-specific size-curve packs.
        4. Category-level proportional packs, if category_group exists.
        5. Channel-level proportional packs for high-volume SKUs.
    """
    candidates: Dict[Tuple[Tuple[str, int], ...], PackCandidate] = {}

    total_by_sku = demand.sum(axis=1).sort_values(ascending=False)

    # 1. Single-SKU packs.
    single_quantities = [1, 2, 3, 4, 5, 6, 8, 10, 12, 16, 20, 24]
    for sku, total in total_by_sku.items():
        for q in single_quantities:
            if q <= MAX_PACK_UNITS and q <= max(1, total):
                add_candidate(candidates, f"single_{sku}_x{q}", {sku: q})

    # Helper table with one row per sku for grouping.
    sku_meta_cols = ["sku_id", "product_id", "size"]
    if "category_group" in forecast.columns:
        sku_meta_cols.append("category_group")
    sku_meta = forecast[sku_meta_cols].drop_duplicates()

    # 2. Same-product size curves based on total demand.
    pack_sizes = [2, 3, 4, 5, 6, 8, 10, 12, 16, 20, 24]
    for pid, sku_rows in sku_meta.groupby("product_id"):
        skus = list(sku_rows["sku_id"])
        product_demand = total_by_sku.reindex(skus).fillna(0)
        product_demand = product_demand[product_demand > 0]
        if product_demand.empty:
            continue

        for pack_size in pack_sizes:
            if pack_size < len(product_demand):
                continue
            content = proportional_integer_pack(
                product_demand,
                target_units=pack_size,
                max_distinct_skus=min(MAX_DISTINCT_SKUS_PER_PACK, len(product_demand)),
            )
            add_candidate(candidates, f"prod_{pid}_curve_{pack_size}", content)

    # 3. Same-product, channel-specific size curves.
    for channel in demand.columns:
        ch_demand = demand[channel]
        for pid, sku_rows in sku_meta.groupby("product_id"):
            skus = list(sku_rows["sku_id"])
            product_ch_demand = ch_demand.reindex(skus).fillna(0)
            product_ch_demand = product_ch_demand[product_ch_demand > 0]
            if product_ch_demand.empty:
                continue

            for pack_size in [4, 6, 8, 10, 12, 16, 20, 24]:
                if pack_size < len(product_ch_demand):
                    continue
                content = proportional_integer_pack(
                    product_ch_demand,
                    target_units=pack_size,
                    max_distinct_skus=min(MAX_DISTINCT_SKUS_PER_PACK, len(product_ch_demand)),
                )
                add_candidate(
                    candidates,
                    f"prod_{pid}_{slugify(channel)}_curve_{pack_size}",
                    content,
                )

    # 4. Category-level packs, if available.
    if "category_group" in sku_meta.columns:
        for cat, sku_rows in sku_meta.groupby("category_group"):
            skus = list(sku_rows["sku_id"])
            cat_demand = total_by_sku.reindex(skus).fillna(0)
            cat_demand = cat_demand[cat_demand > 0].sort_values(ascending=False)
            if cat_demand.empty:
                continue

            for pack_size in [8, 10, 12, 16, 20, 24]:
                content = proportional_integer_pack(
                    cat_demand,
                    target_units=pack_size,
                    max_distinct_skus=min(MAX_DISTINCT_SKUS_PER_PACK, len(cat_demand)),
                )
                add_candidate(candidates, f"cat_{slugify(cat)}_{pack_size}", content)

    # 5. Channel-level proportional packs for high-volume SKUs.
    # These are useful when a channel has a stable broad assortment pattern.
    for channel in demand.columns:
        ch_demand = demand[channel].sort_values(ascending=False)
        ch_demand = ch_demand[ch_demand > 0]
        if ch_demand.empty:
            continue

        for pack_size in [8, 10, 12, 16, 20, 24]:
            content = proportional_integer_pack(
                ch_demand,
                target_units=pack_size,
                max_distinct_skus=MAX_DISTINCT_SKUS_PER_PACK,
            )
            add_candidate(candidates, f"channel_{slugify(channel)}_{pack_size}", content)

    # Sort candidates by a simple usefulness score:
    # larger packs first, then fewer distinct SKUs for operational simplicity.
    packs = list(candidates.values())
    packs.sort(key=lambda p: (-p.total_units, p.distinct_skus, p.name))

    if len(packs) > max_candidate_packs:
        packs = packs[:max_candidate_packs]

    print(f"Generated {len(packs):,} candidate pack types.")
    return packs


# ---------------------------------------------------------------------------
# Optimization model
# ---------------------------------------------------------------------------

def solve_pack_model(
    demand: pd.DataFrame,
    packs: List[PackCandidate],
    unit_costs: Dict[str, float],
    max_pack_types: int | None,
    allow_shortage: bool,
    time_limit: int,
    mip_gap: float,
    threads: int | None,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Build and solve the MIP.

    Returns:
        packs_used_df
        allocation_df
        assortment_df
        cost_summary_df
    """
    skus = list(demand.index)
    channels = list(demand.columns)

    pack_ids = [f"P{idx + 1:03d}" for idx in range(len(packs))]

    # Sparse lookup: sku -> list of (pack_index, quantity)
    sku_to_pack_terms: Dict[str, List[Tuple[int, int]]] = {sku: [] for sku in skus}
    for p_idx, pack in enumerate(packs):
        for sku, qty in pack.content:
            if sku in sku_to_pack_terms:
                sku_to_pack_terms[sku].append((p_idx, qty))

    # Check candidate coverage.
    uncovered = [sku for sku, terms in sku_to_pack_terms.items() if not terms]
    if uncovered:
        raise ValueError(
            f"{len(uncovered)} SKUs have forecasted demand but no candidate pack coverage. "
            f"Examples: {uncovered[:10]}"
        )

    model = gp.Model("prepack_optimization")

    if time_limit:
        model.Params.TimeLimit = time_limit
    if mip_gap is not None:
        model.Params.MIPGap = mip_gap
    if threads:
        model.Params.Threads = threads

    # Decision variables.
    y = model.addVars(len(packs), vtype=GRB.BINARY, name="use_pack")

    x = model.addVars(
        len(packs),
        len(channels),
        vtype=GRB.INTEGER,
        lb=0,
        name="allocate",
    )

    over = model.addVars(
        len(skus),
        len(channels),
        vtype=GRB.CONTINUOUS,
        lb=0.0,
        name="overstock",
    )

    if allow_shortage:
        under = model.addVars(
            len(skus),
            len(channels),
            vtype=GRB.CONTINUOUS,
            lb=0.0,
            name="shortage",
        )
    else:
        under = None

    # Link allocation to pack-use binary.
    # A safe big-M is the maximum number of packs needed if this pack type alone
    # supplied the highest total channel demand.
    max_channel_total = int(demand.sum(axis=0).max())
    for p_idx, pack in enumerate(packs):
        m_p = max(1, math.ceil(max_channel_total / max(1, pack.total_units)) + 1)
        for c_idx in range(len(channels)):
            model.addConstr(x[p_idx, c_idx] <= m_p * y[p_idx], name=f"link_{p_idx}_{c_idx}")

    # Optional submitted-solution complexity limit.
    if max_pack_types is not None:
        model.addConstr(gp.quicksum(y[p_idx] for p_idx in range(len(packs))) <= max_pack_types,
                        name="max_pack_types")

    # Demand balance per SKU/channel.
    shipped_expr = {}
    for s_idx, sku in enumerate(skus):
        terms_for_sku = sku_to_pack_terms[sku]
        for c_idx, channel in enumerate(channels):
            shipped = gp.quicksum(qty * x[p_idx, c_idx] for p_idx, qty in terms_for_sku)
            shipped_expr[(sku, channel)] = shipped
            d = int(demand.loc[sku, channel])

            if allow_shortage:
                model.addConstr(
                    shipped - over[s_idx, c_idx] + under[s_idx, c_idx] == d,
                    name=f"balance_{sku}_{channel}",
                )
            else:
                model.addConstr(
                    shipped - over[s_idx, c_idx] == d,
                    name=f"balance_{sku}_{channel}",
                )

    # Objective components.
    setup_cost = PACK_CREATION_COST * gp.quicksum(y[p_idx] for p_idx in range(len(packs)))

    handling_cost = HANDLING_COST_PER_PACK * gp.quicksum(
        x[p_idx, c_idx]
        for p_idx in range(len(packs))
        for c_idx in range(len(channels))
    )

    capital_cost = COST_OF_CAPITAL * gp.quicksum(
        float(unit_costs.get(sku, 1.0)) * over[s_idx, c_idx]
        for s_idx, sku in enumerate(skus)
        for c_idx in range(len(channels))
    )

    if allow_shortage:
        shortage_cost = SHORTAGE_PENALTY_PER_UNIT * gp.quicksum(
            under[s_idx, c_idx]
            for s_idx in range(len(skus))
            for c_idx in range(len(channels))
        )
    else:
        shortage_cost = 0

    model.setObjective(setup_cost + handling_cost + capital_cost + shortage_cost, GRB.MINIMIZE)

    model.optimize()

    if model.Status not in {GRB.OPTIMAL, GRB.TIME_LIMIT, GRB.SUBOPTIMAL}:
        raise RuntimeError(f"Gurobi ended with status {model.Status}; no usable solution found.")

    # Extract selected packs.
    selected_pidx = [p_idx for p_idx in range(len(packs)) if y[p_idx].X > 0.5]

    selected_pack_ids = {p_idx: f"P{rank + 1:03d}" for rank, p_idx in enumerate(selected_pidx)}

    pack_rows = []
    for p_idx in selected_pidx:
        row = {"pack_id": selected_pack_ids[p_idx], "candidate_name": packs[p_idx].name}
        row.update({sku: 0 for sku in skus})
        for sku, qty in packs[p_idx].content:
            if sku in skus:
                row[sku] = qty
        pack_rows.append(row)

    packs_used_df = pd.DataFrame(pack_rows)

    # Extract pack allocation.
    allocation_rows = []
    for p_idx in selected_pidx:
        row = {"pack_id": selected_pack_ids[p_idx]}
        any_used = False
        for c_idx, channel in enumerate(channels):
            value = int(round(x[p_idx, c_idx].X))
            row[channel] = value
            any_used = any_used or value > 0
        if any_used:
            allocation_rows.append(row)

    allocation_df = pd.DataFrame(allocation_rows)

    # Build assortment shipped by the chosen solution.
    assortment = pd.DataFrame(0, index=skus, columns=channels, dtype=int)
    for p_idx in selected_pidx:
        for c_idx, channel in enumerate(channels):
            packs_allocated = int(round(x[p_idx, c_idx].X))
            if packs_allocated == 0:
                continue
            for sku, qty in packs[p_idx].content:
                if sku in assortment.index:
                    assortment.loc[sku, channel] += qty * packs_allocated

    assortment_df = assortment.reset_index().rename(columns={"index": "sku_id"})
    demand_df = demand.reset_index().rename(columns={"index": "sku_id"})

    # Long-form diagnostics.
    diag_rows = []
    total_over_units = 0
    total_under_units = 0
    total_capital_cost = 0.0

    for sku in skus:
        for channel in channels:
            d = int(demand.loc[sku, channel])
            shipped = int(assortment.loc[sku, channel])
            over_units = max(0, shipped - d)
            under_units = max(0, d - shipped)
            total_over_units += over_units
            total_under_units += under_units
            total_capital_cost += over_units * float(unit_costs.get(sku, 1.0)) * COST_OF_CAPITAL

            diag_rows.append({
                "sku_id": sku,
                "channel_id": channel,
                "forecast_units": d,
                "shipped_units": shipped,
                "overstock_units": over_units,
                "shortage_units": under_units,
                "unit_cost": float(unit_costs.get(sku, 1.0)),
                "capital_cost": over_units * float(unit_costs.get(sku, 1.0)) * COST_OF_CAPITAL,
            })

    diagnostics_df = pd.DataFrame(diag_rows)

    used_pack_count = len(selected_pidx)
    allocated_pack_count = int(allocation_df.drop(columns=["pack_id"]).to_numpy().sum()) if not allocation_df.empty else 0
    total_units_shipped = int(assortment.to_numpy().sum())
    total_forecast_units = int(demand.to_numpy().sum())

    setup_cost_val = used_pack_count * PACK_CREATION_COST
    handling_cost_val = allocated_pack_count * HANDLING_COST_PER_PACK
    shortage_cost_val = total_under_units * SHORTAGE_PENALTY_PER_UNIT if allow_shortage else 0.0
    total_cost_val = setup_cost_val + handling_cost_val + total_capital_cost + shortage_cost_val

    baseline_handling = total_forecast_units * HANDLING_COST_PER_PACK
    baseline_total = baseline_handling
    estimated_savings = baseline_total - total_cost_val

    cost_summary_df = pd.DataFrame([
        {"metric": "objective_value", "value": float(model.ObjVal)},
        {"metric": "used_pack_types", "value": used_pack_count},
        {"metric": "allocated_packs_total", "value": allocated_pack_count},
        {"metric": "forecast_units_total", "value": total_forecast_units},
        {"metric": "shipped_units_total", "value": total_units_shipped},
        {"metric": "overstock_units_total", "value": total_over_units},
        {"metric": "shortage_units_total", "value": total_under_units},
        {"metric": "setup_cost", "value": setup_cost_val},
        {"metric": "handling_cost", "value": handling_cost_val},
        {"metric": "capital_cost", "value": total_capital_cost},
        {"metric": "shortage_cost", "value": shortage_cost_val},
        {"metric": "total_optimized_cost", "value": total_cost_val},
        {"metric": "baseline_no_pack_handling_cost", "value": baseline_handling},
        {"metric": "estimated_savings_vs_unit_handling_only", "value": estimated_savings},
        {"metric": "mip_gap", "value": float(model.MIPGap) if model.SolCount else np.nan},
        {"metric": "runtime_seconds", "value": float(model.Runtime)},
    ])

    return packs_used_df, allocation_df, assortment_df, cost_summary_df, diagnostics_df, demand_df


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def write_solution_workbook(
    output_path: str,
    packs_df: pd.DataFrame,
    allocation_df: pd.DataFrame,
    demand_skus: List[str],
    template_path: str | None = None,
) -> None:
    """
    Write the solution file.

    If a template exists, this tries to fill the Packs and PackAllocation sheets.
    If not, it creates a workbook with exactly those two sheets.
    """
    from openpyxl import Workbook, load_workbook

    output_path = str(output_path)

    if template_path and Path(template_path).exists():
        wb = load_workbook(template_path)
        if "Packs" not in wb.sheetnames or "PackAllocation" not in wb.sheetnames:
            raise ValueError("Template must contain sheets named 'Packs' and 'PackAllocation'.")

        ws_packs = wb["Packs"]
        ws_alloc = wb["PackAllocation"]

        # Detect SKU headers in Packs sheet: assume first row with several non-empty cells.
        # If this fails, the script falls back to recreating the two sheets.
        try:
            header_row = None
            max_non_empty = 0
            for r in range(1, min(20, ws_packs.max_row) + 1):
                non_empty = sum(1 for cell in ws_packs[r] if cell.value not in (None, ""))
                if non_empty > max_non_empty:
                    max_non_empty = non_empty
                    header_row = r

            if header_row is None or max_non_empty < 2:
                raise ValueError("Could not detect pack header row.")

            sku_headers = []
            sku_cols = []
            for cell in ws_packs[header_row]:
                if cell.column == 1:
                    continue
                if cell.value not in (None, ""):
                    sku_headers.append(str(cell.value))
                    sku_cols.append(cell.column)

            pack_start_row = header_row + 1

            # Clear and fill pack matrix.
            for i, (_, pack_row) in enumerate(packs_df.iterrows()):
                excel_row = pack_start_row + i
                ws_packs.cell(excel_row, 1).value = pack_row["pack_id"]
                for sku, col in zip(sku_headers, sku_cols):
                    ws_packs.cell(excel_row, col).value = int(pack_row.get(sku, 0))

            # Fill unused rows with zero if template contains more rows.
            for excel_row in range(pack_start_row + len(packs_df), ws_packs.max_row + 1):
                for col in sku_cols:
                    ws_packs.cell(excel_row, col).value = 0

            # Allocation sheet: detect channel header row similarly.
            header_row_a = None
            max_non_empty_a = 0
            for r in range(1, min(20, ws_alloc.max_row) + 1):
                non_empty = sum(1 for cell in ws_alloc[r] if cell.value not in (None, ""))
                if non_empty > max_non_empty_a:
                    max_non_empty_a = non_empty
                    header_row_a = r

            if header_row_a is None or max_non_empty_a < 2:
                raise ValueError("Could not detect allocation header row.")

            channel_headers = []
            channel_cols = []
            for cell in ws_alloc[header_row_a]:
                if cell.column == 1:
                    continue
                if cell.value not in (None, ""):
                    channel_headers.append(str(cell.value))
                    channel_cols.append(cell.column)

            alloc_start_row = header_row_a + 1
            allocation_lookup = allocation_df.set_index("pack_id") if not allocation_df.empty else pd.DataFrame()

            for i, (_, pack_row) in enumerate(packs_df.iterrows()):
                pack_id = pack_row["pack_id"]
                excel_row = alloc_start_row + i
                ws_alloc.cell(excel_row, 1).value = pack_id
                for channel, col in zip(channel_headers, channel_cols):
                    value = 0
                    if not allocation_lookup.empty and pack_id in allocation_lookup.index:
                        value = int(allocation_lookup.loc[pack_id].get(channel, 0))
                    ws_alloc.cell(excel_row, col).value = value

            for excel_row in range(alloc_start_row + len(packs_df), ws_alloc.max_row + 1):
                for col in channel_cols:
                    ws_alloc.cell(excel_row, col).value = 0

            wb.save(output_path)
            print(f"Filled template solution workbook: {output_path}")
            return

        except Exception as exc:
            print(f"WARNING: Could not safely fill template ({exc}). Creating clean workbook instead.")

    # Create clean workbook.
    wb = Workbook()
    ws_packs = wb.active
    ws_packs.title = "Packs"
    ws_alloc = wb.create_sheet("PackAllocation")

    sku_cols = [c for c in packs_df.columns if c not in {"pack_id", "candidate_name"}]

    ws_packs.append(["pack_id"] + sku_cols)
    for _, row in packs_df.iterrows():
        ws_packs.append([row["pack_id"]] + [int(row.get(sku, 0)) for sku in sku_cols])

    channels = [c for c in allocation_df.columns if c != "pack_id"]
    ws_alloc.append(["pack_id"] + channels)
    for _, row in allocation_df.iterrows():
        ws_alloc.append([row["pack_id"]] + [int(row.get(ch, 0)) for ch in channels])

    wb.save(output_path)
    print(f"Created solution workbook: {output_path}")


def export_outputs(
    output_dir: str,
    packs_df: pd.DataFrame,
    allocation_df: pd.DataFrame,
    assortment_df: pd.DataFrame,
    cost_summary_df: pd.DataFrame,
    diagnostics_df: pd.DataFrame,
    demand_df: pd.DataFrame,
    template_path: str | None,
) -> None:
    """Save all useful optimization outputs."""
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    packs_df.to_csv(out / "optimalisation_packs_gurobi.csv", index=False)
    allocation_df.to_csv(out / "optimalisation_pack_allocation_gurobi.csv", index=False)
    assortment_df.to_csv(out / "optimalisation_assortment_gurobi.csv", index=False)
    cost_summary_df.to_csv(out / "optimalisation_cost_summary_gurobi.csv", index=False)
    diagnostics_df.to_csv(out / "optimalisation_diagnostics_gurobi.csv", index=False)
    demand_df.to_csv(out / "optimalisation_forecast_demand_used_gurobi.csv", index=False)

    write_solution_workbook(
        output_path=str(out / "optimalisation_solution.xlsx"),
        packs_df=packs_df,
        allocation_df=allocation_df,
        demand_skus=[c for c in packs_df.columns if c not in {"pack_id", "candidate_name"}],
        template_path=template_path,
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Optimize pre-pack contents and allocation with Gurobi.")

    parser.add_argument(
        "--forecast",
        default="outputs/master_forecast_2026.csv",
        help="Path to SKU/channel forecast CSV from run_forecasting.py.",
    )
    parser.add_argument(
        "--products",
        default="data/PPP_stu_products.xlsx",
        help="Path to product master file with costs.",
    )
    parser.add_argument(
        "--template",
        default="data/PPP_solutionFile_2026.xlsx",
        help="Optional official solution template workbook.",
    )
    parser.add_argument(
        "--output-dir",
        default="outputs",
        help="Directory where optimization outputs are written.",
    )
    parser.add_argument(
        "--max-candidate-packs",
        type=int,
        default=MAX_CANDIDATE_PACKS,
        help="Maximum number of generated candidate pack types to keep.",
    )
    parser.add_argument(
        "--max-pack-types",
        type=int,
        default=None,
        help=(
            "Maximum number of pack types allowed in the solution. "
            "Default: number of forecasted SKUs."
        ),
    )
    parser.add_argument(
        "--no-shortage",
        action="store_true",
        default=None,
        help="Force every SKU/channel forecast to be covered. May be harder to solve.",
    )
    parser.add_argument(
        "--time-limit",
        type=int,
        default=TIME_LIMIT,
        help="Gurobi time limit in seconds.",
    )
    parser.add_argument(
        "--mip-gap",
        type=float,
        default=MIP_GAP,
        help="Relative MIP gap target.",
    )
    parser.add_argument(
        "--threads",
        type=int,
        default=None,
        help="Number of Gurobi threads. Default lets Gurobi decide.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()

    forecast = load_forecast(args.forecast)
    demand = build_demand_table(forecast)
    unit_costs = load_unit_costs(args.products, forecast)

    if demand.empty:
        raise ValueError("No positive forecast demand found. Check the forecast input file.")

    print("=" * 72)
    print("PRE-PACK OPTIMIZATION")
    print("=" * 72)
    print(f"Forecast file       : {args.forecast}")
    print(f"SKUs with demand    : {len(demand.index):,}")
    print(f"Channels            : {len(demand.columns):,}")
    print(f"Forecast units total: {int(demand.to_numpy().sum()):,}")
    print("=" * 72)

    packs = generate_candidate_packs(
        forecast=forecast,
        demand=demand,
        max_candidate_packs=args.max_candidate_packs,
    )

    max_pack_types = args.max_pack_types
    if max_pack_types is None and USE_AT_MOST_ONE_PACK_PER_SKU_BY_DEFAULT:
        max_pack_types = len(demand.index)

    if args.no_shortage is None:
        args.no_shortage = NO_SHORTAGE_BY_DEFAULT

    print(f"Max selected pack types: {max_pack_types if max_pack_types is not None else 'unlimited'}")
    print(f"Shortage allowed       : {not args.no_shortage}")
    print("=" * 72)

    packs_df, allocation_df, assortment_df, cost_summary_df, diagnostics_df, demand_df = solve_pack_model(
        demand=demand,
        packs=packs,
        unit_costs=unit_costs,
        max_pack_types=max_pack_types,
        allow_shortage=not args.no_shortage,
        time_limit=args.time_limit,
        mip_gap=args.mip_gap,
        threads=args.threads,
    )

    template_path = args.template if Path(args.template).exists() else None

    #export_outputs(
    #    output_dir=args.output_dir,
    #    packs_df=packs_df,
    #    allocation_df=allocation_df,
    #    assortment_df=assortment_df,
    #    cost_summary_df=cost_summary_df,
    #    diagnostics_df=diagnostics_df,
    #    demand_df=demand_df,
    #    template_path=template_path,
    #)

    print("\nCOST SUMMARY")
    print(cost_summary_df.to_string(index=False))
    print("\nDone.")


if __name__ == "__main__":
    main()
